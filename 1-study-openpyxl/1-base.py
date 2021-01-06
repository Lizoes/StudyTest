import openpyxl

# 打开文件,r把后面当成原生字符，\不做转义
excel = openpyxl.load_workbook(r"C:\Users\li.lin\Downloads\学校账单批次模板(21).xlsx")
# 创建一个Excel
# excel_new = openpyxl.Workbook("py_create_excel.xlsx")
print("------------------excel----------------------")
# 获取所有sheet的name：['学校账单批次模板', 'dict_data']
print(excel.sheetnames)
# 获取当前激活的sheet：<Worksheet "学校账单批次模板">
print(excel.active)
# 获取所有的sheet：[<Worksheet "学校账单批次模板">, <Worksheet "dict_data">]
print(excel.worksheets)
# 是否只读：False
print(excel.read_only)
# 编码方式：utf-8
print(excel.encoding)
# 运行结果最后三行属于该属性
print(excel.properties)

print("-------------------sheet---------------------")
sheet = excel.worksheets[0]
sheet1 = excel.get_sheet_by_name("xx")
# sheet标题：学校账单批次模板
print(sheet.title)
# 行、列数
print(sheet.max_row)
print(sheet.min_row)
print(sheet.max_column)
print(sheet.min_column)
# 行：<generator object Worksheet._cells_by_row at 0x03A176B8>
print(sheet.rows)
# 列：<generator object Worksheet._cells_by_col at 0x03A176B8>
print(sheet.columns)
# 整个sheet:<generator object Worksheet.values at 0x038E76B8>
print(sheet.values)


print("-------------------cell---------------------")
# 获取cell对象、以及对应的值（当合并单元格时只有最左边的才有值）
cell = sheet.cell(2, 2)
cell1 = sheet['A2']
print(cell1.value, type(cell1))
