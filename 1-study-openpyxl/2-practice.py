import openpyxl
import datetime
import os
import re


class OperaExcel():

    def __init__(self, path):
        excel = openpyxl.load_workbook(path)
        self.excel = excel
        self.bindSheet = None

    def bind_sheet(self, sheet_name=None):
        """Open the given filename and return the workbook

        :param sheet_name: the path to open or a file-like object
        :type sheet_name: string

        :rtype: :class:

        """
        if sheet_name in self.excel.sheetnames:
            self.bindSheet = self.excel[sheet_name]
        elif sheet_name is None:
            self.bindSheet = self.excel.sheetnames[0]
        else:
            self.bindSheet = None

    def get_value(self, row, col):
        msg = None
        if 0 < row <= self.bindSheet.max_row and 0 < col <= self.bindSheet.max_column:
            value = self.bindSheet.cell(row=row, column=col).value
        else:
            value = None
            msg = "error number: row or col"
        return value, msg

    def set_value(self, row, col, value):
        msg = None
        if 0 < row <= self.bindSheet.max_row and 0 < col <= self.bindSheet.max_column:
            self.bindSheet.cell(row=row, column=col).value = value
            msg = "OK"
        else:
            msg = "error number: row or col"
        return msg

    def save_excel(self, filename):
        self.excel.save(filename)

    def school_bill_batch(self, itemNum):
        rowOfTableTitle = 3
        itemOfCol = 0
        nameOfCol = 2
        for col in self.bindSheet.max_column:
            if "缴费项目" in self.get_value(rowOfTableTitle, col):
                itemOfCol = col

        for row in range(rowOfTableTitle+1, self.bindSheet.max_row+1):      # 遍历每一行
            for n in itemNum:                   # 创建itemNum个缴费项目和金额
                itemName = self.get_value(row=row, col=2)
                self.set_value(row=row, col=itemOfCol, value=itemName)

        now = datetime.datetime.now().strftime("%Y%m%d%H%M")
        self.set_value(1, 1, "2020账单批次")
        self.save_excel("学校账单批次" + now + ".xlsx")


def remove_excel(path):
    """
        remove file where in the directory path and filename start with '学校账单批次'
    """
    files = os.listdir(path)
    remove_files = []
    for filename in files:
        if re.match(r"学校账单批次*", filename):
            file_path = path + "\\" + filename
            remove_files.append(file_path)
            os.remove(file_path)
    print("delete files:", remove_files)
    # for root, dirs, files in os.walk("D:\Code\Python\StudyTest", topdown=True):
    #     print(root, dirs, files)
    #     break


if __name__ == "__main__":
    opera = OperaExcel(r"D:\Code\Python\StudyTest\excel\学校账单批次模板 (10).xlsx")
    opera.bind_sheet("学校账单批次模板")
    opera.school_bill_batch()

    # remove_excel(os.getcwd())

